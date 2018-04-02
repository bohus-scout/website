from openpyxl import load_workbook
import io

def load_list(filename):
    file_str = io.open(filename, 'r', encoding="utf8").read()

    array = file_str.split("\n")
    newarray = []

    for item in array:
        newarray.append(item.split(" "))
    
    newarray.pop()

    return newarray

def write_to_workbook(array):
    wb = load_workbook('wb.xlsx')
    ws = wb['Ledare']

    i = 0
    while i < len(array):
        item = array[i]
        ws = wb[item[2]]

        row = str(ws.max_row + 1)
        ws['A' + row].value = item[0]
        ws['B' + row].value = item[1]
        
        wb.save('wb.xlsx')
        i += 1



arr = load_list('database.txt')
write_to_workbook(arr)



